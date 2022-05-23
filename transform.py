import tkinter
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import os
from decimal import *
import openpyxl
from openpyxl import Workbook

class CriticalPoint:
    def __init__ (self, name, cptype, x, y, z, rho, d2rho, G, V, l1, l2, l3):
        self.name = name
        self.cptype = cptype
        self.x = x
        self.y = y
        self.z = z
        self.rho = rho
        self.d2rho = d2rho
        self.G = G
        self.V = V
        self.l1 = l1
        self.l2 = l2
        self.l3 = l3

    def get_rho(self):
        return self.rho
    def get_d2rho(self):
        return self.d2rho
    def get_name(self):
        return self.name


    def get_G(self):
        return self.G
    def get_V(self):
        return self.V
    def get_l1(self):
        return self.l1
    def get_l2(self):
        return self.l2
    def get_l3(self):
        return self.l3


    def get_x(self):
        return self.x
    def get_y(self):
        return self.y
    def get_z(self):
        return self.z

    def get_xyz(self):
        coord = [self.x, self.y, self.z]
        return coord
    def get_EML(self):
        EML_energy = str(round((Decimal(self.V) * Decimal(313.7545)), 4))[1:]
        return EML_energy
    def get_H(self):
        total_energy = str(round(Decimal(self.G) + Decimal(self.V), 4))
        return total_energy

    def get_type(self):
        return self.cptype
    
def get_EML_sum():
        for CP in CPlist:
            if CP.get_type == '(3,-1)':
                EML_sum = EML_sum + CP.get_EML
        return EML_sum

def main():
    global fileType
    global CPlist
    global filename
    global F1
    global E1
    global E2
    global E3
    global E4
    global basins
    global CPs
    global CheckBox1
    global CheckBox2
    self = Tk()

    filename = None
    CPlist = []

    # creating a notebook widget to display different tabs with different properties
    tabs = tkinter.ttk.Notebook()
    tabs.grid(row=0, column=0, sticky='ewns', padx = 3, pady=3)

    page1 = ttk.Frame(tabs)   # first page for operations concerning the transformation of existing files
    page2 = ttk.Frame(tabs)   # second page for creating various input files
    tabs.add(page1, text='    Transform     ') # adding text to tabs 
    tabs.add(page2, text='     Create       ')

    self.title('Transformation Wizard')
    self.iconbitmap(os.path.dirname(os.path.realpath(__file__)) + '\\' + 'Wizard.ico')

    F1 = LabelFrame(page1, text="Select File",  padx = 3, pady = 3, width=320) # height=240, width=320
    F1.grid(row=0, column=0, columnspan = 3, pady = 5, sticky = 'n')

    browse = Button(F1, text="Browse...", width=15, command=openfile)
    browse.grid(row=0, column=1, padx=1, pady=3, sticky='e')

    E1 = Entry(F1, text="E1", width = 45)
    E1.grid(row=0, column=0, padx=1, pady=3, sticky="ew")

    E2 = Text(page1, width = 80, height = 25)
    E2.config(font=("Courier", 8))
    E2.grid(row=5, column=0, columnspan = 2, padx=3, pady=10, sticky="s")

    E3 = Entry(F1, text="E3")
    E3.grid(row=4, column=0, padx=1, pady=3, sticky = 'ew')

    E4 = Entry(F1, text='E4')
    E4.grid(row=1, column = 0, padx=1, pady=3, sticky = 'ew')

    CPs = IntVar()
    basins = IntVar()

    CheckBox1 = Checkbutton(F1, text = 'Atomic Basins', variable = basins, onvalue = 1, offvalue = 0)

    CheckBox2 = Checkbutton(F1, text = 'Critical Points', variable = CPs, onvalue = 1, offvalue = 0)

    B1 = Button(F1, text="Go!", width=15, command=go)
    B1.grid(row=1, column=1, padx=1, pady=3, sticky = 'e')

    B2 = Button(F1, text="Clear", width=15, command=cancel)
    B2.grid(row=2, column=1, padx=1, pady=3, sticky="e")

    B3 = Button(F1, text = 'Terminate', width=15, command=self.destroy)
    B3.grid(row=3, column=1, padx=1, pady=3, sticky = 'e')

    B4 = Button(F1, text='Output files directory\n Default is the same', width = 20, command=output)
    B4.grid(row=4, column=1, padx=1, pady=3, sticky = 'e')

    S = Scrollbar(page1)
    S.config(command = E2.yview)
    S.grid(row=5, column=2, sticky="nws")
    E2.config(yscrollcommand=S.set)

    self.mainloop()

def openfile():
    global CheckBox1
    global CheckBox2
    global basins
    global CPs
    global filename
    global fileType
    global F1
    filename = filedialog.askopenfilename(filetypes = [('WinXPRO out file', '.out'), ('TOPOND out file', '.outp'), ('All files', '.*')])
    
    E1.delete(0, END)
    E2.delete(1.0, END)
    if ('trho' in filename.split('\\')[-1]) or ('tlap' in filename.split('\\')[-1]) \
        or ('winxpro' in filename.split('\\')[-1]) or (('xfac' or 'XFAC') in filename.split('\\')[-1]):
        E1.insert(0, filename)
    else:
        E1.insert(0,filename) # 'This is not a suitable file!'
        E2.insert(1.0, filename + ' is not a suitable file!')   

    if 'winxpro' not in filename:
        CheckBox1.grid_remove()
        CheckBox2.grid_remove()

    if 'trho' in os.path.basename(filename):
        fileType = 'Topological analysis of Electron density'
        
        E2.delete(1.0, END)
        E4.delete(0, END)
        E4.insert(0,fileType)
 
        CheckBox1.grid_remove()
        CheckBox2.grid_remove()

    if 'tlap' in os.path.basename(filename):
        fileType = 'Topological analysis of Lalplacian'

        E2.delete(1.0, END)
        E4.delete(0, END)
        E4.insert(0, fileType)

        CheckBox1.grid_remove()
        CheckBox2.grid_remove()

    if 'winxpro' in os.path.basename(filename):
        fileType = 'Winxpro out file'

        E2.delete(1.0, END)
        E4.delete(0, END)
        E4.insert(0, fileType)

        CheckBox1.grid(row=3, column=0, sticky = 'ew')
        CheckBox2.grid(row=2, column=0, sticky = 'ew')  

    if ('XFAC' or 'xfac') in os.path.basename(filename):
        fileType = 'Theoretical static structure factors'

        E2.delete(1.0, END)
        E4.delete(0, END)
        E4.insert(0, fileType)

        CheckBox1.grid_remove()
        CheckBox2.grid_remove()

def go():
    global CPs
    global basins
    global dir_name
    global CPlist
    global F1
    global E2
    global E4
    global filename

    CPlist = []

    if filename == None:
        E1.delete(0, END)
        E2.delete(1.0, END)
        E1.insert(0,'Chose a file first!')    
        E2.insert(1.0,'Chose a file first!')
    else:
        fileCount = 0
        fileNames = []
        if os.path.basename(filename).split('.')[1] == '.outp' or os.path.basename(filename).split('.')[1] == '.out':
            fileNames.append(str(os.path.basename(filename)))
            fileCount += 1

        with open(filename) as f:
            fContent = f.readlines()
            print('file is opened: ' + os.path.basename(filename))
            f.close()

        if 'trho' in os.path.basename(filename):
            fileType = 'trho'
            E2.delete(1.0, END)
        if 'tlap' in os.path.basename(filename):
            fileType = 'tlap'
            E2.delete(1.0, END)
        if 'winxpro' in os.path.basename(filename):
            fileType = 'winxpro'
            E2.delete(1.0, END)
        if ('XFAC' or 'xfac') in os.path.basename(filename):
            fileType = 'xfac'
            E2.delete(1.0, END)

        newLine = ''
        newContent = ''
        cpNumber = 0

        # handling the extraction of data from files
        if fileType == 'winxpro':
            if CPs.get() == 1:
                # Looking for the right block of data by lines with special keywords. Everything else is omitted.
                i = 0
                for index, line in enumerate(fContent):
                    if ('The active property is f =  RHO' in fContent[index]):
                        if ('f =  RHO' in fContent[index + 1]):
                            rho_block_start = index + 6
                            print('start is found')
                    if 'h(r) =' in fContent[index] and '-----------' in fContent[index + 3]:
                        rho_block_end = index + 1
                        print('end is found')

                fContent = fContent[rho_block_start:rho_block_end]  

                # deviding the content into 2 parts: the table and the list of CPs
                for index, line in enumerate(fContent):
                    if 'Summary' in line:
                        part1 = fContent[:index]
                        part2 = fContent[index + 1:]

                # deviding the 2nd part into the blocks of each CP
                start_block = 0
                end_block = 0
                blocks = []
                part2_m = []
                for index, line in enumerate(part2):
                    if '(3,-1)' in line:
                        start_block = index
                        blocks.append(start_block)

                # getting the list if lists for each cp from under the table
                for index, i in enumerate(blocks):
                    if i != blocks[-1]:
                        part2_m.append(part2[blocks[index]:blocks[index + 1]])
                    else:
                        part2_m.append(part2[blocks[index]:])


                # for each cp in part1 (table) and in part2_m(list of properties) construct a class CriticalPoint()
                # and append it to the list of CPs CPlist
                CPlist = []
                for index, item in enumerate(part1):
                    newCP = CriticalPoint('CP ' + part1[index].split()[0] + ' (' + part1[index].split()[2] + ', ' + part1[index].split()[5] +')',
                    part1[index].split()[1],
                    part1[index].split()[9], # Distance between the two bonded atoms
                    part1[index].split()[10], # Distance from the 1st atom to the CP
                    part1[index].split()[11], # Distance from the 2nd atom to the CP
                    str(round((Decimal(part2_m[index][1].split()[2]) / Decimal(0.529177**3)), 4)), # 12 **3
                    str(round((Decimal(part2_m[index][2].split()[2]) / Decimal(0.529177**5)), 4)), # 22 **5
                    part2_m[index][3].split()[2], # 32
                    part2_m[index][4].split()[2], # 42
                    str(round((Decimal(part2_m[index][1].split()[6]) / Decimal(0.529177**5)), 4)), # 16 **5
                    str(round((Decimal(part2_m[index][2].split()[6]) / Decimal(0.529177**5)), 4)), # 26 **5
                    str(round((Decimal(part2_m[index][3].split()[6]) / Decimal(0.529177**5)), 4))) # 36 **5

                    CPlist.append(newCP)
            if basins.get() == 1:
                basins_props_number = 0
                basins_props = []
                integration_blocks = []
                i = 0

                # separating the integration data from different tables
                for index, line in enumerate(fContent[i:]):
                    if 'Number of atoms in the asymmetric unit:' in line:
                        atoms_number = int(line.split()[-1])
                    if ' === INTEGRATED PROPERTY WITHIN ATOMIC BASINS ===' in line:
                        new_basins = []
                        basins_block_start = index + 5
                        basins_block_end = basins_block_start + atoms_number
                        basins_props_number += 1 # rho, d2rho, ELECTRON, volume etc
                        basins_props.append(fContent[index + 3].split()[4])
                        i = basins_block_end
                        # need to save the number of total blocks, integrated property, start and end indices.
                        new_basins.append(fContent[index + 3].split()[4]) # the name of the integrated property
                        new_basins.append(basins_block_start) # starting index of the integration block
                        new_basins.append(basins_block_end) # ending index
                        integration_blocks.append(new_basins) # a lis of all the block of property integration over atomic basins  
        if fileType == 'xfac':
            hklStart = 0
            hklEnd = 0
            hkl_column_widht = 4
            f_column_widht = 10
            f_number_widht = 7
            for index, line in enumerate(fContent):
                if 'REAL PART' in line:
                    hklStart = index + 2
                    print('start is found')
                if 'ERROR' in line:
                    hklEnd = index - 1
                    print('end is found')
            fContent = fContent[hklStart:hklEnd]
            # changing the file string by string
            for string in fContent:
                n = 1
                m = 1
                k = 1
                v = 1
                g = 1
                y = 0
                if len(string.split()[0]) < hkl_column_widht:
                    n = hkl_column_widht - len(string.split()[0])
                if len(string.split()[1]) < hkl_column_widht:
                    m = hkl_column_widht - len(string.split()[1])
                if len(string.split()[2]) < hkl_column_widht:
                    k = hkl_column_widht - len(string.split()[2])
            
                if len(str(Decimal(string.split()[-2]))) < f_number_widht:
                    y = f_number_widht - len(str(Decimal(string.split()[-2])))

                if len(str(Decimal(string.split()[-2])) + '0'*y) < f_column_widht:
                    v = f_column_widht - len(str(Decimal(string.split()[-2])) + '0'*y)

                new_string = ' '*n + string.split()[0] + ' '*m + string.split()[1] + ' '*k + string.split()[2] +\
                             ' '*v + str(Decimal(string.split()[-2])) + '0'*y +' 1 \n'
                
                if Decimal(new_string.split()[-2]) < 0.000001:
                    new_string = new_string.replace(new_string.split()[-2], '  0.00000')
                newContent = newContent + new_string
        for index, line in enumerate(fContent):    
            if ('CP N.' in line) and ('CP TYPE' in fContent[index + 3]) and fileType =='trho':
                    # print(line)
                    cpNumber = fContent[index].split()[2]
                    try:
                        newCP = CriticalPoint('CP' + cpNumber,
                        fContent[index + 3].split()[3],
                        fContent[index + 5].split()[5],
                        fContent[index + 5].split()[6],
                        fContent[index + 5].split()[7],
                        str(round(Decimal(fContent[index + 6].split()[3]), 4)),
                        str(round(Decimal(fContent[index + 6].split()[5]), 4)),
                        fContent[index + 7].split()[5],
                        fContent[index + 8].split()[3],
                        str(round(Decimal(fContent[index + 30].split()[5]), 4)),
                        str(round(Decimal(fContent[index + 30].split()[6]), 4)),
                        str(round(Decimal(fContent[index + 30].split()[7]), 4)))
                        
                        CPlist.append(newCP)
                        
                        if Decimal(newCP.x).number_class() == "-Normal":
                            n = 1
                        else:
                            n = 2
                        if Decimal(newCP.y).number_class() == "-Normal":
                            m = 1
                        else:
                            m = 2
                        if Decimal(newCP.z).number_class() == "-Normal":
                            k = 1
                        else:
                            k = 2
                        if len(newCP.d2rho) < 12:
                            l = 12 - len(newCP.d2rho)
                        if len(newCP.get_H()) < 12:
                            s = 12 - len(newCP.get_H())
                        
                        newLine = '\n DUMY' + ' '*n + newCP.get_x() + ' '*m + newCP.get_y() + ' '*k + newCP.get_z() + ' CP' +\
                                cpNumber + ' 0 ! '
                        
                        newContent = newContent + newLine
                    except (IndexError, TypeError):
                        print('Check CP ' + cpNumber + '!')
                        E2.insert(END, 'Check CP ' + cpNumber + '!')
            if ('CP N.' in line) and ('CP TYPE' in fContent[index + 3]) and fileType == 'tlap':
                    
                    print(line)
                    cpNumber = fContent[index].split()[2]
                
                    newCP = CriticalPoint('CP' + cpNumber + '_' + fContent[index].split()[8] + fContent[index].split()[7] + ' ',
                    fContent[index + 3].split()[3],
                    fContent[index + 5].split()[5],
                    fContent[index + 5].split()[6],
                    str(round(Decimal(fContent[index + 5].split()[7]), 4)), # 5 7
                    str(round(Decimal(fContent[index + 6].split()[3]), 4)), # 6 3
                    str(round(Decimal(fContent[index + 6].split()[5]), 4)), # 6 5 
                    fContent[index + 7].split()[5],
                    fContent[index + 8].split()[3],
                    str(round(Decimal(fContent[index + 13].split()[5]), 4)), # 13 5
                    str(round(Decimal(fContent[index + 13].split()[6]), 4)), # 13 6
                    str(round(Decimal(fContent[index + 13].split()[7]), 4))) # 13 7
                    
                    
                    CPlist.append(newCP)

                    if Decimal(newCP.x).number_class() == "-Normal":
                        n = 1
                    else:
                        n = 2
                    if Decimal(newCP.y).number_class() == "-Normal":
                        m = 1
                    else:
                        m = 2
                    if Decimal(newCP.z).number_class() == "-Normal":
                        k = 1
                    else:
                        k = 2
                    if len(newCP.d2rho) < 12:
                        l = 12 - len(newCP.d2rho)
                    if len(newCP.get_H()) < 12:
                        s = 12 - len(newCP.get_H())

                    newLine = '\n DUMY' + ' '*n + newCP.get_x() + ' '*m + newCP.get_y() + ' '*k + newCP.get_z() +\
                               ' ' + newCP.get_name() + ' 0 ! '
                                       
                    newContent = newContent + newLine
            
        # writing and saving output files
        if fileType == 'trho' or fileType == 'tlap':
            try:
                dummy_file = open(dir_name + '\\' + os.path.basename(filename) + '_DUMMY_' + fileType + '.txt', 'w')
            except NameError:
                dir_name = os.path.dirname(filename)
            
            dummy_file = open(dir_name + '\\' + os.path.basename(filename) + '_DUMMY_' + fileType + '.txt', 'w')
            dummy_file.write(newContent)
            dummy_file.close

            if os.path.exists(dir_name + '\\' + os.path.basename(filename).split('.')[0] +'.txt'):
                filepath = dir_name + '/' + os.path.basename(filename).split('.')[0] +'.txt'
                E2.delete(1.0, END)
                E2.insert(1.0, 'Job is done: \n' + filepath )

            wb = Workbook()
            wb.active.title = 'Crystal ' + fileType + ' critical points'
            ws = wb.active

            ws.column_dimensions['A'].width = 9
            ws.column_dimensions['B'].width = 7
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 15
            ws.column_dimensions['K'].width = 15
            ws.column_dimensions['L'].width = 15
            ws.column_dimensions['M'].width = 15

            #name, cptype, x, y, z, rho, d2rho, G, V, l1, l2, l3
            ws['A1'] = 'CP_name'
            ws['B1'] = 'CP_type'
            ws['C1'] = 'x'
            ws['D1'] = 'y'
            ws['E1'] = 'z'
            ws['F1'] = 'rho'
            ws['G1'] = 'd2rho'
            ws['H1'] = 'G'
            ws['I1'] = 'V'
            ws['J1'] = 'l1'
            ws['K1'] = 'l2'
            ws['L1'] = 'l3'

            if fileType == 'trho':
                ws['M1'] = 'EML'

            for indexRow, CP in enumerate(CPlist):
                indexColumn = 0
                for CP_property, property_value in vars(CP).items():
                    indexColumn += 1
                    ws.cell(row = indexRow + 2, column = indexColumn, value = property_value)
                    if Decimal(CP.d2rho) > 0 and fileType == 'trho':
                        ws.cell(row = indexRow + 2, column = 13, value = CP.get_EML() )
                    else:
                        ws.cell(row = indexRow + 2, column = 13, value = None )

            ws.insert_cols(10)
            ws.cell(row=1, column=10, value='H')
            for cell in range(len(CPlist)):
                ws.cell(row = cell+2, column = 10, value = CPlist[cell].get_H())

            if fileType == 'trho':
                ws.cell(row=1, column=ws.max_column + 1, value='EML from G')
                EMLG_col = ws.max_column
                for cell in range(len(CPlist)):
                    if float(CPlist[cell].get_d2rho()) > 0:
                        ws.cell(row = cell+2, column = EMLG_col, value = str(round(Decimal(CPlist[cell].get_G()) * Decimal(0.429) * Decimal(313.7545), 4)))

            wb.save(dir_name + '\\' + filename.split('/')[-1].split('.')[-2] + '.xlsx')

            if os.path.exists(dir_name + '\\' + os.path.basename(filename).split('.')[0] +'.xlsx'):
                filepath = dir_name + '/' + os.path.basename(filename).split('.')[0] +'.xlsx'
                E2.insert(END, '\n' + filepath )
        elif fileType == 'winxpro':
            if CPs.get() == 1:
                wbcp = Workbook()
                wbcp.active.title = 'WinXPRO electron density CPs'
                ws = wbcp.active

                ws['A1'] = 'CP_name'
                ws['B1'] = 'CP_type'
                ws['C1'] = 'Rij'
                ws['D1'] = 'd1'
                ws['E1'] = 'd2'
                ws['F1'] = 'rho'
                ws['G1'] = 'd2rho'
                ws['H1'] = 'G'
                ws['I1'] = 'V'
                ws['J1'] = 'l1'
                ws['K1'] = 'l2'
                ws['L1'] = 'l3'
                ws['M1'] = 'EML'

                ws.column_dimensions['A'].width = 17
                ws.column_dimensions['B'].width = 8
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 15
                ws.column_dimensions['I'].width = 15
                ws.column_dimensions['J'].width = 15
                ws.column_dimensions['K'].width = 15
                ws.column_dimensions['L'].width = 15
                ws.column_dimensions['M'].width = 15
                ws.column_dimensions['O'].width = 12

                for indexRow, CP in enumerate(CPlist):
                    indexColumn = 0
                    for CP_property, property_value in vars(CP).items():
                        indexColumn += 1
                        ws.cell(row = indexRow + 2, column = indexColumn, value = property_value)
                        if Decimal(CP.d2rho) > 0:
                            ws.cell(row = indexRow + 2, column = 13, value = CP.get_EML() )
                        else:
                            ws.cell(row = indexRow + 2, column = 13, value = None )
                
                ws.insert_cols(10)

                #calculating total energy H and adding it after V
                ws.cell(row=1, column=10, value='H')
                for cell in range(len(CPlist)):
                    ws.cell(row = cell+2, column = 10, value = CPlist[cell].get_H())
                wbcp.save(dir_name + '\\' + os.path.basename(filename).split('.')[0] +'.xlsx')

                # calculating EML from G and adding it to the last column
                ws.cell(row=1, column=ws.max_column + 1, value='EML from G') 
                for cell in range(len(CPlist)):
                    if float(CPlist[cell].get_d2rho()) > 0:
                        ws.cell(row = cell+2, column = ws.max_column, value = str(round(Decimal(CPlist[cell].get_G()) * Decimal(0.429) * Decimal(313.7545), 4)))
                
                wbcp.save(dir_name + '\\' + os.path.basename(filename).split('.')[0] +'.xlsx')

                if os.path.exists(dir_name + '\\' + os.path.basename(filename).split('.')[0] +'.xlsx'):
                    filepath = dir_name + '\\' + os.path.basename(filename).split('.')[0] +'.xlsx'
                    E2.delete(1.0, END)
                    E2.insert(1.0, 'Job is done: \n' + filepath )
            if basins.get() == 1:
                wb = Workbook()
                wb.active.title = 'WinXPRO integration'
                ws = wb.active

                ws['A1'] = 'Atom name'
                for index, prop in enumerate(basins_props):
                    ws.cell(row=1, column=index+2, value=prop)

                for atom in range(1, atoms_number + 1):
                    ws.cell(row=atom + 1, column = 1, value = fContent[atom + 30].split()[0])
                    
                for number, prop in enumerate(integration_blocks): # getting the integrated property
                    for index, line in enumerate(fContent[(integration_blocks[number][1]):(integration_blocks[number][2])]): # going line by line to extract the value for each atom
                        try:
                            ws.cell(row = index + 2, column = number + 2, value=line.split()[4])
                        except IndexError:
                            E2.insert('end', 'something is wrong!')
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
     
                wb.save(dir_name + '\\' + os.path.basename(filename).split('.')[0] +'_basins.xlsx')

                if os.path.exists(dir_name + '\\' + os.path.basename(filename).split('.')[0] +'_basins.xlsx'):
                    filepath = dir_name + '\\' + os.path.basename(filename).split('.')[0] +'_basins.xlsx'
                    #E2.delete(1.0, END)
                    E2.insert('end', 'Job is done: \n' + filepath )
        elif fileType == 'xfac':
            try:
                dummy_file = open(dir_name + '\\' + os.path.basename(filename) + '.hkl', 'w')
            except NameError:
                dir_name = os.path.dirname(filename)
            
            dummy_file = open(dir_name + '\\' + os.path.basename(filename) +'.hkl', 'w')
            dummy_file.write(newContent)
            dummy_file.close
            if os.path.exists(dir_name + '\\' + os.path.basename(filename).split('.')[0] +'.hkl'):
                filepath = dir_name + '/' + os.path.basename(filename).split('.')[0] +'.hkl'
                E2.delete(1.0, END)
                E2.insert(1.0, 'Job is done: \n' + filepath )
          
def cancel():
    global E2
    global E1
    global CheckBox1
    global CheckBox2
    global filename
    filename = None
    E1.delete(0, END)
    E2.delete(1.0, END)
    E4.delete(0, END)
    CheckBox1.grid_remove()
    CheckBox2.grid_remove()

def output():
    global dir_name
    global E3
    dir_name = filedialog.askdirectory()
    E3.delete(0, END)
    E3.insert(0, dir_name)

main()